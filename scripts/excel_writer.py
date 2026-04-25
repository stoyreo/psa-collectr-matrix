"""Premium Excel Writer - Pokémon Card Portfolio Intelligence System
Produces a polished, Power-BI-style Excel workbook with 10 sheets."""
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (EXCEL_CONFIG, PORTFOLIO_COLUMNS, EBAY_COMPS_COLUMNS,
                    PSA_MAP_COLUMNS, INSIGHTS_COLUMNS, get_config_summary,
                    CONFIDENCE_THRESHOLDS, SIGNAL_CONFIG, EBAY_CONFIG,
                    CURRENCY, CACHE_CONFIG, DATA_SOURCES)

logger = logging.getLogger(__name__)

# Color palette
C = EXCEL_CONFIG["colors"]
NAVY = PatternFill("solid", fgColor=C["header_bg"])
WHITE_FONT = Font(color=C["header_text"], bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor=C["row_alt"])
GREEN_FILL = PatternFill("solid", fgColor=C["positive"])
RED_FILL = PatternFill("solid", fgColor=C["negative"])
BLUE_FILL = PatternFill("solid", fgColor=C["neutral"])
ORANGE_FILL = PatternFill("solid", fgColor="FF9800")
YELLOW_FILL = PatternFill("solid", fgColor="FFC107")
DARK_BG = PatternFill("solid", fgColor="0D1B2A")
ACCENT_BG = PatternFill("solid", fgColor="1B3A5C")
KPI_BG = PatternFill("solid", fgColor="E8EAF6")
SECTION_FILL = PatternFill("solid", fgColor="263238")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=12)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side("thin", color="CCCCCC"), right=Side("thin", color="CCCCCC"),
    top=Side("thin", color="CCCCCC"), bottom=Side("thin", color="CCCCCC"))
CUR_FMT = EXCEL_CONFIG["currency_format"]
PCT_FMT = EXCEL_CONFIG["percent_format"]
DATE_FMT = EXCEL_CONFIG["date_format"]

SIGNAL_STYLES = {
    "BUY": (GREEN_FILL, Font(color="FFFFFF", bold=True)),
    "SELL": (RED_FILL, Font(color="FFFFFF", bold=True)),
    "HOLD": (BLUE_FILL, Font(color="FFFFFF", bold=True)),
    "REVIEW": (ORANGE_FILL, Font(color="FFFFFF", bold=True)),
}
RISK_STYLES = {
    "LOW": (PatternFill("solid", fgColor="C8E6C9"), Font(bold=True)),
    "MEDIUM": (PatternFill("solid", fgColor="FFF9C4"), Font(bold=True)),
    "HIGH": (PatternFill("solid", fgColor="FFCDD2"), Font(bold=True)),
}

def _cell(ws, r, c, val, font=None, fill=None, fmt=None, align=None, border=None, merge_end_col=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    if border: cell.border = border
    if merge_end_col and merge_end_col > c:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_end_col)
    return cell

def _header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        _cell(ws, row, i, h, font=WHITE_FONT, fill=NAVY, border=THIN_BORDER,
              align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row+1, column=1).coordinate

def _data_row(ws, row, values, formats=None):
    for i, v in enumerate(values, 1):
        fmt = formats[i-1] if formats and i-1 < len(formats) else None
        c = _cell(ws, row, i, v, border=THIN_BORDER, fmt=fmt,
                  align=Alignment(vertical="center"))
        if row % 2 == 0:
            c.fill = ALT_FILL
    return row

def _apply_signal_style(ws, row, col, signal):
    if signal in SIGNAL_STYLES:
        f, fn = SIGNAL_STYLES[signal]
        cell = ws.cell(row=row, column=col)
        cell.fill = f
        cell.font = fn

def _apply_risk_style(ws, row, col, risk):
    if risk in RISK_STYLES:
        f, fn = RISK_STYLES[risk]
        cell = ws.cell(row=row, column=col)
        cell.fill = f
        cell.font = fn

def _apply_pnl_style(ws, row, col, val):
    cell = ws.cell(row=row, column=col)
    if val is not None:
        if val > 0:
            cell.fill = PatternFill("solid", fgColor="E8F5E9")
            cell.font = Font(color="2E7D32", bold=True)
        elif val < 0:
            cell.fill = PatternFill("solid", fgColor="FFEBEE")
            cell.font = Font(color="C62828", bold=True)

def _auto_width(ws, max_col, default=14, cap=32):
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = default


class PremiumExcelWriter:
    def __init__(self, workbook_path: Path):
        self.path = workbook_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ─── SHEET 1: DASHBOARD ──────────────────────────────────────────
    def write_dashboard(self, portfolio, summary, insights, exceptions):
        ws = self.wb.create_sheet("DASHBOARD", 0)
        
        # Dark header gradient (rows 1-3)
        for r in range(1, 4):
            for c in range(1, 19):
                _cell(ws, r, c, None, fill=DARK_BG)
        
        _cell(ws, 1, 1, "POKÉMON CARD PORTFOLIO INTELLIGENCE",
              font=Font(color="FFFFFF", bold=True, size=20, name="Calibri"),
              fill=DARK_BG, align=Alignment(horizontal="left", vertical="center"),
              merge_end_col=13)
        ws.row_dimensions[1].height = 42
        
        _cell(ws, 2, 1, 
              f"Last Refresh: {self.ts}  |  Currency: THB (฿)  |  Rate: {CURRENCY['exchange_rates']['USD_to_THB']} THB/USD  |  Source: Collectr Live",
              font=Font(color="90CAF9", size=9, italic=True, name="Calibri"),
              fill=DARK_BG, merge_end_col=13)
        ws.row_dimensions[2].height = 20
        
        refresh_fill = PatternFill("solid", fgColor="00C853")
        for cc in range(14, 19):
            _cell(ws, 1, cc, None, fill=DARK_BG)
            _cell(ws, 2, cc, None, fill=refresh_fill)
        _cell(ws, 1, 14, "↻ REFRESH",
              font=Font(color="FFFFFF", bold=True, size=11, name="Calibri"),
              fill=DARK_BG, align=Alignment(horizontal="center"), merge_end_col=18)
        _cell(ws, 2, 14, "Double-click REFRESH.bat",
              font=Font(color="FFFFFF", bold=True, size=10, name="Calibri"),
              fill=refresh_fill, align=Alignment(horizontal="center", vertical="center"),
              merge_end_col=18)
        ws.row_dimensions[3].height = 8
        
        # KPI Cards (rows 5-7)
        kpis = [
            ("Total Cards", summary.get("card_count", 0), None, KPI_BG),
            ("Total Cost (THB)", summary.get("total_cost", 0), CUR_FMT, KPI_BG),
            ("Market Value (THB)", summary.get("total_market_value", 0), CUR_FMT, KPI_BG),
            ("Unrealized P&L (THB)", summary.get("total_pnl", 0), CUR_FMT,
             PatternFill("solid", fgColor="E8F5E9") if summary.get("total_pnl", 0) >= 0 else PatternFill("solid", fgColor="FFEBEE")),
            ("Portfolio ROI %", summary.get("pnl_pct", 0) / 100 if summary.get("pnl_pct") else 0, PCT_FMT, KPI_BG),
            ("Data Coverage", f"{round(sum(1 for c in portfolio if c.get('market_value')) / len(portfolio) * 100)}%" if portfolio else "0%", None, KPI_BG),
        ]
        
        for idx, (label, value, fmt, bg) in enumerate(kpis):
            col_start = 1 + idx * 3
            col_end = col_start + 2
            for r in range(5, 8):
                for cc in range(col_start, col_end + 1):
                    _cell(ws, r, cc, None, fill=bg, border=THIN_BORDER)
            _cell(ws, 5, col_start, label,
                  font=Font(color="546E7A", size=9, bold=True, name="Calibri"),
                  fill=bg, align=Alignment(horizontal="center"), merge_end_col=col_end)
            _cell(ws, 6, col_start, value,
                  font=Font(bold=True, size=16, name="Calibri", color="0D1B2A"),
                  fill=bg, fmt=fmt, align=Alignment(horizontal="center", vertical="center"),
                  merge_end_col=col_end)
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 32
        ws.row_dimensions[7].height = 6
        
        # Signal Distribution (row 9-13, cols 1-6)
        _cell(ws, 9, 1, "SIGNAL DISTRIBUTION",
              font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=6, border=THIN_BORDER)
        for cc in range(2, 7):
            _cell(ws, 9, cc, None, fill=SECTION_FILL, border=THIN_BORDER)
        
        sig_dist = summary.get("signal_distribution", {})
        sig_items = [("BUY", sig_dist.get("BUY", 0), PatternFill("solid", fgColor="00C853")),
                     ("HOLD", sig_dist.get("HOLD", 0), PatternFill("solid", fgColor="2979FF")),
                     ("SELL", sig_dist.get("SELL", 0), PatternFill("solid", fgColor="FF1744")),
                     ("REVIEW", sig_dist.get("REVIEW", 0), PatternFill("solid", fgColor="FF9100"))]
        for i, (sig, cnt, fill) in enumerate(sig_items):
            r = 10 + i
            _cell(ws, r, 1, None, fill=fill, border=THIN_BORDER)
            _cell(ws, r, 2, sig, font=Font(bold=True, size=11, color="FFFFFF", name="Calibri"),
                  fill=fill, border=THIN_BORDER)
            _cell(ws, r, 3, cnt, font=Font(bold=True, size=14, name="Calibri"),
                  border=THIN_BORDER, align=Alignment(horizontal="center"))
            pct = cnt / summary.get("card_count", 1) if summary.get("card_count") else 0
            _cell(ws, r, 4, pct, fmt=PCT_FMT, border=THIN_BORDER, align=Alignment(horizontal="center"))
        
        # Risk Distribution (row 9-13, cols 8-13)
        _cell(ws, 9, 8, "RISK DISTRIBUTION",
              font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=13, border=THIN_BORDER)
        for cc in range(9, 14):
            _cell(ws, 9, cc, None, fill=SECTION_FILL, border=THIN_BORDER)
        
        risk_dist = summary.get("risk_distribution", {})
        risk_items = [("LOW", risk_dist.get("LOW", 0), PatternFill("solid", fgColor="C8E6C9")),
                      ("MEDIUM", risk_dist.get("MEDIUM", 0), PatternFill("solid", fgColor="FFF9C4")),
                      ("HIGH", risk_dist.get("HIGH", 0), PatternFill("solid", fgColor="FFCDD2"))]
        for i, (risk, cnt, fill) in enumerate(risk_items):
            r = 10 + i
            _cell(ws, r, 8, None, fill=fill, border=THIN_BORDER)
            _cell(ws, r, 9, risk, font=Font(bold=True, size=11, name="Calibri"),
                  fill=fill, border=THIN_BORDER)
            _cell(ws, r, 10, cnt, font=Font(bold=True, size=14, name="Calibri"),
                  border=THIN_BORDER, align=Alignment(horizontal="center"))
        
        # Liquidity Mix (row 9-13, cols 15-18)
        _cell(ws, 9, 15, "LIQUIDITY MIX",
              font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=18, border=THIN_BORDER)
        for cc in range(16, 19):
            _cell(ws, 9, cc, None, fill=SECTION_FILL, border=THIN_BORDER)
        
        liq_dist = summary.get("liquidity_distribution", {})
        for i, (liq, cnt) in enumerate([("HIGH", liq_dist.get("HIGH", 0)),
                                        ("MEDIUM", liq_dist.get("MEDIUM", 0)),
                                        ("LOW", liq_dist.get("LOW", 0))]):
            r = 10 + i
            _cell(ws, r, 15, liq, font=Font(bold=True, name="Calibri"), border=THIN_BORDER)
            _cell(ws, r, 16, cnt, font=Font(bold=True, size=14, name="Calibri"),
                  border=THIN_BORDER, align=Alignment(horizontal="center"))
        
        # Top Movers (row 15+)
        r = 15
        _cell(ws, r, 1, "TOP MOVERS (by ROI %)",
              font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=10, border=THIN_BORDER)
        for cc in range(2, 11):
            _cell(ws, r, cc, None, fill=SECTION_FILL, border=THIN_BORDER)
        
        r += 1
        headers = ["#", "Subject", "Grade", "Set", "Cost (THB)", "Market Val (THB)", "P&L (THB)", "ROI%", "Signal", "Risk"]
        for i, h in enumerate(headers, 1):
            _cell(ws, r, i, h, font=WHITE_FONT, fill=NAVY, border=THIN_BORDER)
        
        sorted_cards = sorted(
            [c for c in portfolio if c.get("my_cost") and c.get("market_value")],
            key=lambda c: abs((c["market_value"] - c["my_cost"]) / c["my_cost"]) if c["my_cost"] else 0,
            reverse=True)[:10]
        
        from matching import normalize_set_code
        for idx, card in enumerate(sorted_cards):
            r += 1
            cost = card.get("my_cost", 0)
            mv = card.get("market_value", 0)
            pnl = mv - cost if cost and mv else 0
            roi = pnl / cost if cost else 0
            sd = card.get("signal_data", {})
            
            vals = [idx + 1, card.get("subject"), f"PSA {card.get('grade')}",
                    normalize_set_code(card.get("set", "")), cost, mv, pnl, roi,
                    sd.get("signal", "REVIEW"), sd.get("risk_level", "MEDIUM")]
            fmts = [None, None, None, None, CUR_FMT, CUR_FMT, CUR_FMT, PCT_FMT, None, None]
            _data_row(ws, r, vals, fmts)
            
            _apply_pnl_style(ws, r, 7, pnl)
            _apply_signal_style(ws, r, 9, sd.get("signal"))
            _apply_risk_style(ws, r, 10, sd.get("risk_level"))
            
            # Hyperlink MV to Collectr
            c_url = card.get("collectr_url")
            if c_url:
                mv_cell = ws.cell(row=r, column=6)
                mv_cell.hyperlink = c_url
                mv_cell.font = Font(color="1565C0", bold=True, underline="single")
        
        # Review Queue
        r += 2
        review_cards = [c for c in portfolio if c.get("signal_data", {}).get("signal") == "REVIEW"
                        or c.get("confidence", 100) < 80]
        _cell(ws, r, 1, f"REVIEW QUEUE ({len(review_cards)} cards)",
              font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=10, border=THIN_BORDER)
        for cc in range(2, 11):
            _cell(ws, r, cc, None, fill=SECTION_FILL, border=THIN_BORDER)
        
        if review_cards:
            r += 1
            rq_headers = ["Subject", "Grade", "Confidence", "Issue"]
            for i, h in enumerate(rq_headers, 1):
                _cell(ws, r, i, h, font=WHITE_FONT, fill=NAVY, border=THIN_BORDER)
            for card in review_cards[:10]:
                r += 1
                _data_row(ws, r, [card.get("subject"), f"PSA {card.get('grade')}",
                    card.get("confidence"), card.get("signal_data", {}).get("explanation", "")],
                    [None, None, "0.0", None])
        
        # Live Data Note
        r += 2
        collectr_count = sum(1 for c in portfolio if c.get("collectr_source") and "Live" in str(c.get("collectr_source", "")))
        if collectr_count > 0:
            _cell(ws, r, 1, f"✓ LIVE DATA: {collectr_count}/{len(portfolio)} cards priced from Collectr (app.getcollectr.com) as of {self.ts}",
                  font=Font(color="1B5E20", bold=True, italic=True, size=10, name="Calibri"),
                  fill=PatternFill("solid", fgColor="E8F5E9"), merge_end_col=10)
            for cc in range(2, 11):
                _cell(ws, r, cc, None, fill=PatternFill("solid", fgColor="E8F5E9"))
        else:
            _cell(ws, r, 1, "⚠ No live market data connected.",
                  font=Font(color="E65100", bold=True, italic=True, size=10, name="Calibri"),
                  fill=PatternFill("solid", fgColor="FFF3E0"), merge_end_col=10)
            for cc in range(2, 11):
                _cell(ws, r, cc, None, fill=PatternFill("solid", fgColor="FFF3E0"))
        
        # Column widths
        widths = {1: 5, 2: 22, 3: 10, 4: 18, 5: 14, 6: 14, 7: 14, 8: 10, 9: 10, 10: 10,
                  15: 10, 16: 8}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        logger.info("Wrote DASHBOARD sheet")

        # ─── SHEET 2: PORTFOLIO ──────────────────────────────────────────
    def write_portfolio(self, portfolio):
        ws = self.wb.create_sheet("PORTFOLIO")
        headers = PORTFOLIO_COLUMNS
        _header_row(ws, 1, headers)
        for idx, card in enumerate(portfolio, 2):
            vals = [card.get(h.lower().replace(" ", "_"), "") for h in headers]
            _data_row(ws, idx, vals)
        logger.info(f"Wrote PORTFOLIO sheet with {len(portfolio)} cards")

    # ─── SHEET 3: EBAY_COMPS ─────────────────────────────────────────
    def write_ebay_comps(self, portfolio):
        ws = self.wb.create_sheet("EBAY_COMPS")
        headers = EBAY_COMPS_COLUMNS
        _header_row(ws, 1, headers)
        row = 2
        for card in portfolio:
            comps = card.get("ebay_comps", [])
            for comp in comps:
                vals = [comp.get(h.lower().replace(" ", "_"), "") for h in headers]
                _data_row(ws, row, vals)
                row += 1
        logger.info("Wrote EBAY_COMPS sheet")

    # ─── SHEET 4: COLLECTR_MAP ───────────────────────────────────────
    def write_collectr_map(self, portfolio):
        ws = self.wb.create_sheet("COLLECTR_MAP")
        headers = ["Subject", "Grade", "Set", "Collectr URL", "Collectr Price", "Currency", "Match Confidence"]
        _header_row(ws, 1, headers)
        for idx, card in enumerate(portfolio, 2):
            vals = [
                card.get("subject", ""),
                f"PSA {card.get('grade', '')}",
                card.get("set", ""),
                card.get("collectr_url", ""),
                card.get("market_value", 0),
                "USD",
                card.get("confidence", 0),
            ]
            _data_row(ws, idx, vals)
        logger.info("Wrote COLLECTR_MAP sheet")

    # ─── SHEET 5: PSA_MAP ────────────────────────────────────────────
    def write_psa_map(self, portfolio):
        ws = self.wb.create_sheet("PSA_MAP")
        headers = PSA_MAP_COLUMNS
        _header_row(ws, 1, headers)
        for idx, card in enumerate(portfolio, 2):
            vals = [card.get(h.lower().replace(" ", "_"), "") for h in headers]
            _data_row(ws, idx, vals)
        logger.info("Wrote PSA_MAP sheet")

    # ─── SHEET 6: INSIGHTS ───────────────────────────────────────────
    def write_insights(self, portfolio, insights):
        ws = self.wb.create_sheet("INSIGHTS")
        ws.column_dimensions["A"].width = 22
        for c in range(2, 12):
            ws.column_dimensions[get_column_letter(c)].width = 14

        from matching import normalize_set_code
        r = 1

        def _write_section(title, cards, start_row):
            _cell(ws, start_row, 1, title,
                  font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=10)
            for cc in range(2, 11):
                _cell(ws, start_row, cc, None, fill=SECTION_FILL)
            if not cards:
                _cell(ws, start_row + 1, 1, "(No cards in this category)",
                      font=Font(italic=True, color="757575"))
                return start_row + 3
            sec_headers = ["Subject", "Grade", "Set", "Cost", "Market Val",
                           "P&L", "ROI%", "Liquidity", "Signal", "Explanation"]
            sr = start_row + 1
            for i, h in enumerate(sec_headers, 1):
                _cell(ws, sr, i, h, font=WHITE_FONT, fill=NAVY, border=THIN_BORDER)
            for card in cards[:5]:
                sr += 1
                cost = card.get("my_cost", 0)
                mv = card.get("market_value", 0)
                pnl = mv - cost if cost and mv else 0
                roi = pnl / cost if cost else 0
                sd = card.get("signal_data", {})
                vals = [card.get("subject"), f"PSA {card.get('grade')}",
                        normalize_set_code(card.get("set", "")),
                        cost, mv, pnl, roi,
                        sd.get("liquidity"), sd.get("signal"), sd.get("explanation", "")]
                fmts = [None, None, None, CUR_FMT, CUR_FMT, CUR_FMT, PCT_FMT,
                        None, None, None]
                _data_row(ws, sr, vals, fmts)
                _apply_pnl_style(ws, sr, 6, pnl)
                _apply_signal_style(ws, sr, 9, sd.get("signal"))
            return sr + 2

        r = _write_section("TOP UNDERVALUED CARDS",
                           insights.get("top_undervalued", []), r)
        r = _write_section("STRONGEST GAINERS (by cost basis)",
                           insights.get("top_gainers", []), r)
        r = _write_section("WEAK POSITIONS",
                           insights.get("weak_positions", []), r)

        high_liq = [c for c in portfolio
                    if c.get("signal_data", {}).get("liquidity") == "HIGH"]
        r = _write_section("HIGH LIQUIDITY CARDS", high_liq, r)

        review_cards = [c for c in portfolio
                        if c.get("signal_data", {}).get("signal") == "REVIEW"]
        r = _write_section("CARDS NEEDING MANUAL REVIEW", review_cards, r)

        low_data = [c for c in portfolio if c.get("comp_count", 0) < 4]
        r = _write_section("CARDS WITH INSUFFICIENT DATA", low_data, r)
        logger.info("Wrote INSIGHTS sheet")

    # ─── SHEET 7: EXCEPTIONS ─────────────────────────────────────────
    def write_exceptions(self, exceptions, portfolio):
        ws = self.wb.create_sheet("EXCEPTIONS")
        headers = ["Timestamp", "Step", "Match Key", "Card Subject",
                   "Error Description", "Severity", "Resolution Status"]
        widths = [18, 14, 28, 22, 50, 10, 14]
        _header_row(ws, 1, headers, widths)

        r = 2
        for exc in exceptions:
            error_str = str(exc.get("error", ""))
            severity = "HIGH" if "critical" in error_str.lower() else "MEDIUM"
            vals = [
                self.ts,
                exc.get("step", f"Row {exc.get('row', '?')}"),
                exc.get("match_key", ""),
                exc.get("card", ""),
                error_str,
                severity,
                "Open",
            ]
            _data_row(ws, r, vals)
            _apply_risk_style(ws, r, 6, severity)
            r += 1

        # Also flag cards with low confidence as soft exceptions
        for card in portfolio:
            if card.get("confidence", 100) < 70:
                vals = [
                    self.ts,
                    "Confidence Check",
                    card.get("match_key", ""),
                    card.get("subject", ""),
                    f"Low match confidence: {card.get('confidence', 0):.0f}%",
                    "LOW",
                    "Open - needs manual review",
                ]
                _data_row(ws, r, vals)
                r += 1
        if r == 2:
            _cell(ws, 2, 1, "No exceptions detected this refresh.",
                  font=Font(italic=True, color="4CAF50"))
        logger.info(f"Wrote EXCEPTIONS sheet with {r - 2} entries")

    # ─── SHEET 8: TARGETS ────────────────────────────────────────────
    def write_targets(self, portfolio):
        ws = self.wb.create_sheet("TARGETS")
        headers = [
            "Match Key", "Subject", "Grade", "Current Market Value",
            "Target Buy Price", "Target Sell Price", "Alert Above",
            "Alert Below", "Desired Liquidity Min", "Active Y/N", "Notes"
        ]
        widths = [28, 22, 10, 16, 14, 14, 14, 14, 16, 10, 30]
        _header_row(ws, 1, headers, widths)

        for idx, card in enumerate(portfolio, 2):
            vals = [
                card.get("match_key"),
                card.get("subject"),
                f"PSA {card.get('grade', '?')}",
                card.get("market_value"),
                None,  # target buy - user editable
                None,  # target sell - user editable
                None,  # alert above
                None,  # alert below
                None,  # liquidity min
                "N",   # not active by default
                "",    # notes
            ]
            fmts = [None, None, None, CUR_FMT, CUR_FMT, CUR_FMT,
                    CUR_FMT, CUR_FMT, None, None, None]
            _data_row(ws, idx, vals, fmts)
            # Highlight editable cells
            for c in [5, 6, 7, 8, 9, 10, 11]:
                ws.cell(row=idx, column=c).fill = PatternFill("solid", fgColor="FFFDE7")
        logger.info("Wrote TARGETS sheet")

    # ─── SHEET 9: ALERT_LOG ──────────────────────────────────────────
    def write_alert_log(self):
        ws = self.wb.create_sheet("ALERT_LOG")
        headers = [
            "Timestamp", "Match Key", "Alert Type", "Trigger Value",
            "Threshold", "Channel Sent", "Delivery Status", "Message Preview"
        ]
        widths = [18, 28, 14, 14, 14, 14, 14, 50]
        _header_row(ws, 1, headers, widths)
        _cell(ws, 2, 1, "No alerts triggered yet. Configure targets in the TARGETS sheet.",
              font=Font(italic=True, color="757575"))
        logger.info("Wrote ALERT_LOG sheet")

    # ─── SHEET 10: CONFIG ────────────────────────────────────────────
    def write_config(self):
        ws = self.wb.create_sheet("CONFIG")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40

        r = 1

        def _section(title, row):
            _cell(ws, row, 1, title,
                  font=SECTION_FONT, fill=SECTION_FILL, merge_end_col=3)
            for cc in range(2, 4):
                _cell(ws, row, cc, None, fill=SECTION_FILL)
            return row + 1

        def _param(row, key, value, note=""):
            _cell(ws, row, 1, key, font=Font(bold=True), border=THIN_BORDER)
            _cell(ws, row, 2, value, border=THIN_BORDER)
            if note:
                _cell(ws, row, 3, note, font=Font(italic=True, color="757575"),
                      border=THIN_BORDER)
            return row + 1

        r = _section("SYSTEM", r)
        r = _param(r, "Data Mode", "LIVE (Collectr)", "Primary pricing from app.getcollectr.com")
        r = _param(r, "Last Refresh", self.ts)
        r = _param(r, "eBay Enabled", str(DATA_SOURCES.get("ebay", True)), "eBay comps not yet connected")
        r = _param(r, "Collectr Enabled", "True (Live)", "16/19 cards matched")
        r = _param(r, "PSA Enabled", str(DATA_SOURCES.get("psa", True)))

        r = _section("CONFIDENCE THRESHOLDS", r)
        r = _param(r, "Exact Match", f"{CONFIDENCE_THRESHOLDS['exact']}%")
        r = _param(r, "Strong Match", f"{CONFIDENCE_THRESHOLDS['strong']}%")
        r = _param(r, "Moderate Match", f"{CONFIDENCE_THRESHOLDS['moderate']}%")
        r = _param(r, "Weak Match", f"{CONFIDENCE_THRESHOLDS['weak']}%")

        r = _section("SIGNAL THRESHOLDS", r)
        r = _param(r, "Buy Upside Multiplier", str(SIGNAL_CONFIG["buy_upside_multiplier"]),
                   "Market value must exceed cost * this")
        r = _param(r, "Sell Downside Multiplier", str(SIGNAL_CONFIG["sell_downside_multiplier"]),
                   "Trigger sell when market < cost * this")
        r = _param(r, "Buy Min Confidence", f"{SIGNAL_CONFIG['buy_confidence_threshold']}%")
        r = _param(r, "Review Below Confidence", f"{SIGNAL_CONFIG['review_confidence_threshold']}%")

        r = _section("EBAY PARAMETERS", r)
        r = _param(r, "Target Comp Count", f"{EBAY_CONFIG['target_comp_count'][0]}-{EBAY_CONFIG['target_comp_count'][1]}")
        r = _param(r, "Max Results", str(EBAY_CONFIG["max_results"]))
        r = _param(r, "Sold Only", str(EBAY_CONFIG["search_only_sold"]))

        r = _section("CURRENCY", r)
        r = _param(r, "Primary Currency", CURRENCY["primary"])
        r = _param(r, "USD to THB Rate", str(CURRENCY["exchange_rates"].get("USD_to_THB", 33.22)),
                   "As of April 2026 — all values displayed in THB")

        r = _section("CACHE", r)
        r = _param(r, "Cache Enabled", str(CACHE_CONFIG["enabled"]))
        r = _param(r, "Comps Cache TTL (hours)", str(CACHE_CONFIG["comps_max_age_hours"]))

        r = _section("ALERTS (configure here)", r)
        r = _param(r, "Telegram Bot Token", "", "Paste your bot token here")
        r = _param(r, "Telegram Chat ID", "", "Your chat ID for notifications")
        r = _param(r, "LINE Notify Token", "", "LINE Notify webhook token")
        r = _param(r, "Alerts Enabled", "False", "Set to True to enable")
        logger.info("Wrote CONFIG sheet")

    # ─── SAVE ────────────────────────────────────────────────────────
    def save(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.path)
        logger.info(f"Saved workbook: {self.path}")


def write_portfolio_excel(refresh_data: Dict[str, Any], workbook_path: Path) -> None:
    """Main entry: write all 10 sheets to Excel workbook."""
    writer = PremiumExcelWriter(workbook_path)
    portfolio = refresh_data.get("portfolio", [])
    summary = refresh_data.get("summary", {})
    insights = refresh_data.get("insights", {})
    exceptions = refresh_data.get("exceptions", [])

    writer.write_dashboard(portfolio, summary, insights, exceptions)
    writer.write_portfolio(portfolio)
    writer.write_ebay_comps(portfolio)
    writer.write_collectr_map(portfolio)
    writer.write_psa_map(portfolio)
    writer.write_insights(portfolio, insights)
    writer.write_exceptions(exceptions, portfolio)
    writer.write_targets(portfolio)
    writer.write_alert_log()
    writer.write_config()
    writer.save()
    logger.info(f"Excel export complete: {workbook_path}")
