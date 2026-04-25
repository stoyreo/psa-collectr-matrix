"""Read/write the Portfolio_Master.xlsx master workbook."""
from __future__ import annotations

import os
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
try:
    from filelock import FileLock
except ImportError:
    # Graceful fallback: filelock not installed → use a no-op stub.
    # Owned-card annotation and read-only access still work; only
    # concurrent writers lose their cross-process lock.
    class FileLock:
        def __init__(self, *a, **kw): pass
        def __enter__(self): return self
        def __exit__(self, *a): return False
        def acquire(self, *a, **kw): return self
        def release(self, *a, **kw): pass

logger = logging.getLogger(__name__)

MASTER_PATH = Path(__file__).parent.parent / "Portfolio_Master.xlsx"
LOCK_PATH = Path(__file__).parent.parent / "Portfolio_Master.xlsx.lock"

CARD_COLUMNS = [  # 33 CSV + 6 new
    "Item Status", "Item", "Cert Number", "Grade Issuer", "Grade",
    "Autograph Grade", "Year", "Set", "Card Number", "Subject", "Variety",
    "Serial", "Category", "My Cost", "PSA Estimate", "Gain/Loss", "My Value",
    "Date Acquired", "Source", "My Notes", "Vault Status", "Vaulted Date",
    "Days Vaulted", "Listing Status", "Listing Date", "Listing Price",
    "Sold Status", "Sold On", "Sold Date", "Sold Price", "Sold Fees",
    "Sold Proceeds", "Payment Date",
    # new columns
    "collectr_url", "collectr_price_thb", "ebay_avg_thb", "ebay_n_comps",
    "image_match_ok", "last_updated",
]


def _get_lock():
    """Get a file lock with 10-second timeout."""
    return FileLock(str(LOCK_PATH), timeout=10)


def ensure_master_exists() -> Path:
    """Create a fresh Portfolio_Master.xlsx with 3 sheets if missing."""
    if MASTER_PATH.exists():
        logger.info(f"Master workbook already exists at {MASTER_PATH}")
        return MASTER_PATH

    logger.info(f"Creating new master workbook at {MASTER_PATH}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Cards
    ws_cards = wb.create_sheet("Cards", 0)
    ws_cards.append(CARD_COLUMNS)

    # Format header: bold navy on white
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_cards[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze row 1
    ws_cards.freeze_panes = "A2"

    # Set column widths
    for col in ws_cards.columns:
        ws_cards.column_dimensions[col[0].column_letter].width = 14

    # Sheet 2: Sync Log
    ws_log = wb.create_sheet("Sync Log", 1)
    ws_log.append(["timestamp", "source", "cert", "action", "note"])
    for cell in ws_log[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws_log.freeze_panes = "A2"
    for col in ws_log.columns:
        ws_log.column_dimensions[col[0].column_letter].width = 20

    # Sheet 3: Meta
    ws_meta = wb.create_sheet("Meta", 2)
    ws_meta.append(["schema_version", "1"])
    ws_meta.append(["created_at", datetime.now().isoformat()])
    ws_meta.append(["last_csv_sync", ""])
    ws_meta.append(["usd_to_thb_rate", "33.22"])

    for row in ws_meta.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 30

    with _get_lock():
        MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(MASTER_PATH))

    logger.info(f"Master workbook created with 3 sheets")
    return MASTER_PATH


def append_card(row: dict, source: str) -> int:
    """Append one row to Cards sheet; log to Sync Log. Returns new row index.
       source ∈ {'csv', 'add-card', 'refresh'}. Uses atomic save (tmp + rename)."""

    with _get_lock():
        wb = load_workbook(str(MASTER_PATH))
        ws_cards = wb["Cards"]

        # Build row in column order
        row_values = []
        for col_name in CARD_COLUMNS:
            row_values.append(row.get(col_name, ""))

        # Append the row
        ws_cards.append(row_values)
        new_row_idx = ws_cards.max_row

        # Log to Sync Log
        cert = row.get("Cert Number", "")
        ws_log = wb["Sync Log"]
        ws_log.append([
            datetime.now().isoformat(),
            source,
            cert,
            "added",
            row.get("My Notes", "")[:50] if row.get("My Notes") else ""
        ])

        # Update Meta sheet
        ws_meta = wb["Meta"]
        ws_meta["B3"].value = datetime.now().isoformat()  # last_csv_sync

        # Atomic save
        tmp_path = MASTER_PATH.parent / f"{MASTER_PATH.name}.tmp"
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(MASTER_PATH))

        logger.info(f"Appended card {cert} from {source} at row {new_row_idx}")
        return new_row_idx


def update_card(cert: str, fields: dict, source: str) -> int:
    """Update the most recent row for this cert. Always log to Sync Log.
       Returns updated row index. If cert not found, raises KeyError."""

    with _get_lock():
        wb = load_workbook(str(MASTER_PATH))
        ws_cards = wb["Cards"]

        # Find the most recent row with this cert (search from bottom)
        cert_col_idx = CARD_COLUMNS.index("Cert Number") + 1  # 1-indexed
        row_idx = None
        for row in range(ws_cards.max_row, 1, -1):
            if str(ws_cards.cell(row, cert_col_idx).value) == str(cert):
                row_idx = row
                break

        if row_idx is None:
            raise KeyError(f"Cert {cert} not found in master workbook")

        # Update the fields
        for col_name, value in fields.items():
            if col_name in CARD_COLUMNS:
                col_idx = CARD_COLUMNS.index(col_name) + 1
                ws_cards.cell(row_idx, col_idx).value = value

        # Always set last_updated
        last_updated_col_idx = CARD_COLUMNS.index("last_updated") + 1
        ws_cards.cell(row_idx, last_updated_col_idx).value = datetime.now().isoformat()

        # Log to Sync Log
        ws_log = wb["Sync Log"]
        ws_log.append([
            datetime.now().isoformat(),
            source,
            cert,
            "updated",
            f"Updated {', '.join(fields.keys())}"[:50]
        ])

        # Atomic save
        tmp_path = MASTER_PATH.parent / f"{MASTER_PATH.name}.tmp"
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(MASTER_PATH))

        logger.info(f"Updated card {cert} from {source} at row {row_idx}")
        return row_idx


def list_cards() -> list[dict]:
    """Return all Cards rows as list of dicts (header-keyed)."""
    with _get_lock():
        wb = load_workbook(str(MASTER_PATH))
        ws_cards = wb["Cards"]

        cards = []
        for row_idx, row in enumerate(ws_cards.iter_rows(min_row=2, values_only=False), start=2):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                col_name = CARD_COLUMNS[col_idx]
                row_dict[col_name] = cell.value
            cards.append(row_dict)

        return cards


def append_sync_log(source: str, cert: str, action: str, note: str = "") -> None:
    """Append a sync log entry."""
    with _get_lock():
        wb = load_workbook(str(MASTER_PATH))
        ws_log = wb["Sync Log"]

        ws_log.append([
            datetime.now().isoformat(),
            source,
            cert,
            action,
            note[:50] if note else ""
        ])

        # Atomic save
        tmp_path = MASTER_PATH.parent / f"{MASTER_PATH.name}.tmp"
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(MASTER_PATH))
